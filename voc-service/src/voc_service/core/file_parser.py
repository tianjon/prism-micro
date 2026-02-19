"""CSV/Excel 文件解析器。"""

import csv
import io
from dataclasses import dataclass

import chardet
from fastapi import UploadFile

from prism_shared.exceptions import AppException


@dataclass
class ParseResult:
    """文件解析结果。"""

    rows: list[dict[str, str]]
    columns: list[str]
    total_rows: int
    file_size_bytes: int
    detected_encoding: str
    detected_format: str


def _detect_encoding(raw_bytes: bytes) -> str:
    """自动检测文件编码。取前 64KB 采样检测，避免大文件全量扫描。"""
    sample = raw_bytes[:65536]
    result = chardet.detect(sample)
    encoding = result.get("encoding")
    if not encoding:
        raise AppException(
            code="VOC_FILE_ENCODING_ERROR",
            message="无法识别文件编码",
            status_code=400,
        )
    # chardet 有时返回 GB2312，统一映射为 GBK（超集）
    normalized = encoding.lower()
    if normalized in ("gb2312", "gbk", "gb18030"):
        return "gbk"
    return normalized


def _parse_csv_bytes(raw_bytes: bytes, *, sample_only: bool, sample_rows: int) -> ParseResult:
    """解析 CSV 字节流。"""
    encoding = _detect_encoding(raw_bytes)
    try:
        text = raw_bytes.decode(encoding)
    except (UnicodeDecodeError, LookupError) as e:
        raise AppException(
            code="VOC_FILE_ENCODING_ERROR",
            message=f"使用 {encoding} 编码解码失败：{e}",
            status_code=400,
        ) from e

    # 去除 BOM
    if text.startswith("\ufeff"):
        text = text[1:]

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise AppException(
            code="VOC_EMPTY_FILE",
            message="CSV 文件为空或缺少表头",
            status_code=400,
        )

    columns = list(reader.fieldnames)
    rows: list[dict[str, str]] = []
    for i, row in enumerate(reader):
        if sample_only and i >= sample_rows:
            break
        rows.append(row)

    # 如果是采样模式，总行数需要继续计数
    total_rows = len(rows)
    if sample_only and total_rows >= sample_rows:
        for _ in reader:
            total_rows += 1

    return ParseResult(
        rows=rows,
        columns=columns,
        total_rows=total_rows,
        file_size_bytes=len(raw_bytes),
        detected_encoding=encoding,
        detected_format="csv",
    )


def _parse_excel_bytes(raw_bytes: bytes, *, sample_only: bool, sample_rows: int) -> ParseResult:
    """解析 Excel 字节流。"""
    from openpyxl import load_workbook

    # read_only=False：部分 Excel 文件（复杂格式/合并单元格）在 read_only 模式下
    # 仅返回首列，导致数据丢失。文件已全部在内存中，无流式读取收益。
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=False)
    try:
        ws = wb.active
        if ws is None:
            raise AppException(
                code="VOC_EMPTY_FILE",
                message="Excel 文件没有活动工作表",
                status_code=400,
            )

        row_iter = ws.iter_rows(values_only=True)
        # 第一行作为表头
        try:
            header_row = next(row_iter)
        except StopIteration as e:
            raise AppException(
                code="VOC_EMPTY_FILE",
                message="Excel 文件为空",
                status_code=400,
            ) from e

        columns = [str(c).strip() if c is not None else f"column_{i}" for i, c in enumerate(header_row)]
        if not any(c for c in columns if not c.startswith("column_")):
            raise AppException(
                code="VOC_EMPTY_FILE",
                message="Excel 文件表头为空",
                status_code=400,
            )

        rows: list[dict[str, str]] = []
        total_rows = 0
        for row_values in row_iter:
            total_rows += 1
            if sample_only and len(rows) >= sample_rows:
                continue  # 继续计数但不添加行
            row_dict = {}
            for j, val in enumerate(row_values):
                if j < len(columns):
                    row_dict[columns[j]] = str(val) if val is not None else ""
            rows.append(row_dict)
    finally:
        wb.close()

    return ParseResult(
        rows=rows,
        columns=columns,
        total_rows=total_rows,
        file_size_bytes=len(raw_bytes),
        detected_encoding="utf-8",
        detected_format="excel",
    )


def _detect_file_type(filename: str) -> str:
    """根据文件名推断文件类型，不合法则抛出异常。"""
    ext = ""
    dot_idx = filename.rfind(".")
    if dot_idx >= 0:
        ext = filename[dot_idx:].lower()

    file_type = _ALLOWED_EXTENSIONS.get(ext)
    if not file_type:
        raise AppException(
            code="VOC_INVALID_FILE_FORMAT",
            message=f"不支持的文件格式 '{ext}'，仅支持 .csv / .xlsx",
            status_code=400,
        )
    return file_type


def _validate_bytes(raw_bytes: bytes, *, max_file_size_bytes: int) -> None:
    """校验文件字节流是否合法。"""
    if not raw_bytes:
        raise AppException(
            code="VOC_EMPTY_FILE",
            message="上传文件为空",
            status_code=400,
        )
    if len(raw_bytes) > max_file_size_bytes:
        raise AppException(
            code="VOC_FILE_TOO_LARGE",
            message=f"文件大小 {len(raw_bytes)} 字节超过限制 {max_file_size_bytes} 字节",
            status_code=400,
        )


_ALLOWED_EXTENSIONS = {
    ".csv": "csv",
    ".xlsx": "excel",
}


def parse_bytes(
    raw_bytes: bytes,
    *,
    filename: str,
    max_file_size_bytes: int = 52_428_800,
    sample_only: bool = False,
    sample_rows: int = 10,
) -> ParseResult:
    """从原始字节解析文件（同步，供后台任务使用）。"""
    file_type = _detect_file_type(filename)
    _validate_bytes(raw_bytes, max_file_size_bytes=max_file_size_bytes)

    if file_type == "csv":
        return _parse_csv_bytes(raw_bytes, sample_only=sample_only, sample_rows=sample_rows)
    return _parse_excel_bytes(raw_bytes, sample_only=sample_only, sample_rows=sample_rows)


async def parse_file(
    file: UploadFile,
    *,
    max_file_size_bytes: int = 52_428_800,
    sample_only: bool = False,
    sample_rows: int = 10,
) -> ParseResult:
    """
    统一入口：根据扩展名分派到 CSV/Excel 解析器。

    Args:
        file: FastAPI UploadFile 对象
        max_file_size_bytes: 最大文件大小限制
        sample_only: 是否只返回采样数据
        sample_rows: 采样行数
    """
    filename = file.filename or ""
    raw_bytes = await file.read()
    return parse_bytes(
        raw_bytes,
        filename=filename,
        max_file_size_bytes=max_file_size_bytes,
        sample_only=sample_only,
        sample_rows=sample_rows,
    )


def compute_column_statistics(
    rows: list[dict[str, str]],
    columns: list[str],
) -> list[dict]:
    """计算各列的统计信息（简易版，用于采样数据）。

    Returns:
        [{"name": "列名", "unique_count": N, "null_count": N, "sample_values": [...]}]
    """
    stats = []
    for col in columns:
        values = [row.get(col, "") for row in rows]
        non_empty = [v for v in values if v.strip()]
        unique_values = set(non_empty)
        null_count = len(values) - len(non_empty)
        sample_values = list(unique_values)[:5]
        stats.append({
            "name": col,
            "unique_count": len(unique_values),
            "null_count": null_count,
            "sample_values": sample_values,
        })
    return stats


def compute_full_column_statistics(raw_bytes: bytes, filename: str) -> list[dict]:
    """使用 pandas 对全量数据计算列统计信息。

    包含：数据类型推断、唯一值数、空值数、高频样本值、数值范围/均值。
    """
    import pandas as pd

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        encoding = _detect_encoding(raw_bytes)
        df = pd.read_csv(io.BytesIO(raw_bytes), encoding=encoding, dtype=str)
    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
    else:
        return []

    total_rows = len(df)
    stats = []

    for col in df.columns:
        series = df[col].fillna("")
        is_empty = series.str.strip() == ""
        null_count = int(is_empty.sum())
        non_empty = series[~is_empty]
        total_count = int(len(non_empty))
        unique_count = int(non_empty.nunique())

        # Top 5 最频繁值
        top_values = non_empty.value_counts().head(5).index.tolist() if total_count > 0 else []
        sample_values = [str(v) for v in top_values]

        # 推断数据类型 + 数值统计
        dtype = "text"
        min_val = None
        max_val = None
        mean_val = None

        if total_count > 0:
            numeric = pd.to_numeric(non_empty, errors="coerce")
            numeric_valid = numeric.dropna()
            numeric_ratio = len(numeric_valid) / total_count
            if numeric_ratio > 0.8:
                dtype = "numeric"
                if len(numeric_valid) > 0:
                    min_val = str(numeric_valid.min())
                    max_val = str(numeric_valid.max())
                    mean_val = f"{numeric_valid.mean():.2f}"

        stats.append({
            "name": str(col),
            "dtype": dtype,
            "total_count": total_count,
            "total_rows": total_rows,
            "unique_count": unique_count,
            "null_count": null_count,
            "sample_values": sample_values,
            "min_value": min_val,
            "max_value": max_val,
            "mean_value": mean_val,
        })

    return stats


def generate_dataframe_statistics(raw_bytes: bytes, filename: str) -> str:
    """生成 pandas DataFrame 统计信息字符串（info + describe），供 LLM 提示词使用。"""
    desc_text, info_text = generate_split_statistics(raw_bytes, filename)
    return (
        f"### DataFrame.info()\n```\n{info_text}```\n\n"
        f"### DataFrame.describe(include='all')\n```\n{desc_text}```"
    )


def generate_split_statistics(raw_bytes: bytes, filename: str) -> tuple[str, str]:
    """分别返回 (describe_text, info_text)，供 V3 模板使用。"""
    import pandas as pd

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        encoding = _detect_encoding(raw_bytes)
        df = pd.read_csv(io.BytesIO(raw_bytes), encoding=encoding)
    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(io.BytesIO(raw_bytes))
    else:
        return ("不支持的文件格式", "不支持的文件格式")

    buf = io.StringIO()
    df.info(buf=buf)
    info_text = buf.getvalue()
    desc_text = df.describe(include="all").to_string()
    return (desc_text, info_text)


def random_sample_rows(
    raw_bytes: bytes,
    filename: str,
    n: int = 5,
    seed: int = 42,
) -> str:
    """随机采样 N 行数据，返回格式化字符串供 LLM 提示词使用。

    使用固定 seed 确保同一文件多次采样结果一致（幂等性）。
    """
    import pandas as pd

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        encoding = _detect_encoding(raw_bytes)
        df = pd.read_csv(io.BytesIO(raw_bytes), encoding=encoding, dtype=str)
    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
    else:
        return "不支持的文件格式"

    actual_n = min(n, len(df))
    sample_df = df.sample(n=actual_n, random_state=seed)
    return sample_df.to_string(index=False)
